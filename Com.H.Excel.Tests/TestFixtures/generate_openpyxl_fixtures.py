# Generates the openpyxl-produced .xlsx files used as regression fixtures.
# Re-run with:  uv run --with openpyxl python generate_openpyxl_fixtures.py
# (run from the TestFixtures folder)
#
# These fixtures pin down the real-world failure mode: a downstream app reading an
# openpyxl-authored xlsx hit a NullReferenceException in the library because the
# read path was not robust to openpyxl's encoding choices. We commit the .xlsx
# outputs so tests don't depend on uv/python at runtime.

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import Cell

HERE = Path(__file__).resolve().parent


def write_basic_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "City"])
    ws.append(["Alice", "Paris"])
    ws.append(["Bob", "Berlin"])
    ws.append(["Charlie", "Tokyo"])
    wb.save(HERE / "openpyxl_basic_strings.xlsx")


def write_mixed_types():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Score", "When", "Active"])
    ws.append(["alice", 42, datetime(2024, 1, 15), True])
    ws.append(["bob", 7, datetime(2024, 6, 30), False])
    wb.save(HERE / "openpyxl_mixed_types.xlsx")


def write_empty_middle_cells():
    # openpyxl normally writes every cell, but assigning None should leave them
    # represented as empty cells in the resulting xml. This exercises the
    # collapsed-column path on real openpyxl output.
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "C", "D"])
    ws.append(["a1", None, "c1", "d1"])
    ws.append(["a2", "b2", None, "d2"])
    wb.save(HERE / "openpyxl_empty_middle.xlsx")


if __name__ == "__main__":
    write_basic_strings()
    write_mixed_types()
    write_empty_middle_cells()
    print("Generated fixtures in", HERE)
